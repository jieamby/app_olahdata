import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.font import Font
from tkcalendar import DateEntry, Calendar
import pandas as pd
from sqlalchemy import create_engine, Column, Integer, String, func
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import datetime

# Setup SQLite Database
DATABASE_URI = 'sqlite:///database.db'
engine = create_engine(DATABASE_URI, echo=True)
Session = sessionmaker(bind=engine)
session = Session()
Base = declarative_base()

# Define the database model
class Record(Base):
    __tablename__ = 'records'
    
    id = Column(Integer, primary_key=True)
    nomor_urut = Column(Integer, nullable=False)
    nomor_akta = Column(String)
    tanggal_akta = Column(String)
    sifat_akta = Column(String)
    nama_debitur = Column(String)
    gender = Column(String)
    status = Column(String)
    nama_perwakilan = Column(String)
    perusahaan_finance = Column(String)
    alamat_perusahaan = Column(String)
    gelar = Column(String)

# Create the table
Base.metadata.create_all(engine)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title('Aplikasi Olah Data')
        self.root.geometry('1000x600')  # Set window size to 1000x600

        # Configure Treeview style
        self.style = ttk.Style()
        self.style.configure('Treeview.Heading', background='#007acc', foreground='green', font=('Arial', 10, 'bold'))
        self.style.configure('Treeview', font=('Arial', 8), rowheight=25)

        # Define row colors
        self.style.configure('evenrow', background='#C9C0BB')  # Light gray for even rows
        self.style.configure('oddrow', background='#ffffff')  # White for odd rows

        # Create the button frame
        self.button_frame = tk.Frame(root)
        self.button_frame.pack(pady=10, fill='x')

        # Upload, Input, Download Buttons
        self.upload_button = tk.Button(self.button_frame, text="Upload Data", command=self.open_upload_form)
        self.upload_button.pack(side='left', padx=5)

        self.input_button = tk.Button(self.button_frame, text="Input Data", command=self.open_input_form)
        self.input_button.pack(side='left', padx=5)

        self.download_button = tk.Button(self.button_frame, text="Download Data", command=self.download_data)
        self.download_button.pack(side='left', padx=5)

        self.delete_button = tk.Button(self.button_frame, text="Hapus Data", command=self.delete_data)
        self.delete_button.pack(side='left', padx=5)

        self.delete_all_button = tk.Button(self.button_frame, text="Hapus Database", command=self.delete_all_data)
        self.delete_all_button.pack(side='right', padx=5)

        # Create search entry and button
        self.search_frame = tk.Frame(root)
        self.search_frame.pack(pady=10, fill='x')

        self.search_entry = tk.Entry(self.search_frame, width=50)
        self.search_entry.pack(side='right', padx=10)

        self.search_button = tk.Button(self.search_frame, text="Search", command=self.filter_data)
        self.search_button.pack(side='right', padx=10)

        # Create Treeview
        self.tree_frame = tk.Frame(root)
        self.tree = ttk.Treeview(self.tree_frame, columns=('Nomor Urut', 'Nomor Akta', 'Tanggal Akta', 'Sifat Akta', 'Nama Debitur', 'Gender', 'Status', 'Nama Perwakilan', 'Perusahaan Finance', 'Alamat Perusahaan', 'Gelar'), show='headings')
        self.tree_frame.pack(expand=True, fill='both')

        # Set column headings and alignments
        self.tree.heading('Nomor Urut', text='No. Urut')
        self.tree.heading('Nomor Akta', text='No. Akta')
        self.tree.heading('Tanggal Akta', text='Tanggal Akta')
        self.tree.heading('Sifat Akta', text='Sifat Akta')
        self.tree.heading('Nama Debitur', text='Debitur')
        self.tree.heading('Gender', text='Gender')
        self.tree.heading('Status', text='Marital')        
        self.tree.heading('Nama Perwakilan', text='Nama Perwakilan')
        self.tree.heading('Perusahaan Finance', text='Perusahaan Finance')
        self.tree.heading('Alamat Perusahaan', text='Alamat Perusahaan')
        self.tree.heading('Gelar', text='Status')

        # Set column widths
        self.tree.column('Nomor Urut', width=20, anchor='center')
        self.tree.column('Nomor Akta', width=20, anchor='center')
        self.tree.column('Tanggal Akta', width=100, anchor='center')
        self.tree.column('Sifat Akta', width=150, anchor='w')
        self.tree.column('Nama Debitur', width=150, anchor='w')
        self.tree.column('Gender', width=50, anchor='w')
        self.tree.column('Status', width=50, anchor='w')
        self.tree.column('Nama Perwakilan', width=150, anchor='w')
        self.tree.column('Perusahaan Finance', width=150, anchor='w')
        self.tree.column('Alamat Perusahaan', width=200, anchor='w')
        self.tree.column('Gelar', width=20, anchor='center')

        # # Add vertical scrollbar
        self.vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=self.vsb.set)

        self.tree.pack(expand=True, fill='both')

        self.display_data()
        self.tree.bind('<Double-1>', self.on_item_double_click)

        # Initialize log file
        self.log_file_path = "log.txt"
        self.log("Application started")

    def log(self, message):
        """Append a log message to the log file with timestamp."""
        with open(self.log_file_path, 'a') as log_file:
            log_file.write(f"{datetime.datetime.now()} - {message}\n")

    def open_upload_form(self):
        self.upload_window = tk.Toplevel(self.root)
        self.upload_window.title("Upload Data")
        self.upload_window.geometry("400x300")  # Set modal size to 400x200

        # Create a frame for each input row to align labels and entries
        input_frame = tk.Frame(self.upload_window)
        input_frame.pack(pady=10)

        # Nomor Urut Pertama
        tk.Label(input_frame, text="Nomor Urut Pertama:").grid(row=0, column=0, padx=10, pady=5, sticky='e')
        self.start_nomor_urut_entry = tk.Entry(input_frame)
        self.start_nomor_urut_entry.grid(row=0, column=1, padx=10, pady=5)

        # Nama Perwakilan
        tk.Label(input_frame, text="Nama Perwakilan:").grid(row=1, column=0, padx=10, pady=5, sticky='e')
        self.nama_perwakilan_entry = tk.Entry(input_frame)
        self.nama_perwakilan_entry.grid(row=1, column=1, padx=10, pady=5)

        # Nama Perusahaan Finance
        tk.Label(input_frame, text="Nama Perusahaan Finance:").grid(row=2, column=0, padx=10, pady=5, sticky='e')
        self.nama_perusahaan_finance_entry = tk.Entry(input_frame)
        self.nama_perusahaan_finance_entry.grid(row=2, column=1, padx=10, pady=5)

        # Alamat Perusahaan Finance
        tk.Label(input_frame, text="Alamat Perusahaan").grid(row=3, column=0, padx=10, pady=5, sticky='e')
        self.alamat_perusahaan_finance = tk.Entry(input_frame)
        self.alamat_perusahaan_finance.grid(row=3, column=1, padx=10, pady=5)
        
        # Upload File Button
        tk.Button(self.upload_window, text="Upload File", command=self.upload_file).pack(pady=10)
    
    def upload_file(self):
        start_nomor_urut = self.start_nomor_urut_entry.get()
        if not start_nomor_urut.isdigit():
            messagebox.showerror("Error", "Nomor Urut harus angka.")
            return
        
        start_nomor_urut = int(start_nomor_urut)

        # Get the value of Nama Perwakilan from the entry
        nama_perwakilan = self.nama_perwakilan_entry.get()
        if not nama_perwakilan:
            messagebox.showerror("Error", "Nama Perwakilan tidak boleh kosong.")
            return

        # Get the value of Nama Perusahaan Finance from the entry
        nama_perusahaan_finance = self.nama_perusahaan_finance_entry.get()
        if not nama_perusahaan_finance:
            messagebox.showerror("Error", "Nama Perusahaan Finance tidak boleh kosong.")
            return

        # Get the value of Nama Perusahaan Finance from the entry
        alamat_perusahaan_finance = self.alamat_perusahaan_finance.get()
        if not alamat_perusahaan_finance:
            messagebox.showerror("Error", "Alamat Perusahaan Finance tidak boleh kosong.")
            return
        
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.process_file(file_path, start_nomor_urut, nama_perwakilan, nama_perusahaan_finance, alamat_perusahaan_finance)
            self.log(f"Data uploaded from file: {file_path}")

    def process_file(self, file_path, start_nomor_urut, nama_perwakilan, nama_perusahaan_finance, alamat_perusahaan_finance):
        try:
            # Read the Excel file
            df = pd.read_excel(file_path)
            
            # Create a list of records to insert into the database
            records = []
            existing_nomor_urut = {record.nomor_urut for record in session.query(Record.nomor_urut).all()}

            current_nomor_urut = start_nomor_urut

            for _, row in df.iterrows():
                nomor_akta = row.get('number', '')
                # Convert date from DD/MM/YYYY to DD Month YYYY
                tanggal_akta = row.get('created_time', '')
                if pd.notna(tanggal_akta):
                    try:
                        tanggal_akta = datetime.datetime.strptime(tanggal_akta, "%d/%m/%Y %H:%M").strftime("%d %B %Y")
                    except ValueError:
                        try:
                            tanggal_akta = datetime.datetime.strptime(tanggal_akta, "%d/%m/%Y").strftime("%d %B %Y")
                        except ValueError:
                            tanggal_akta = ''
                else:
                    tanggal_akta = ''
                            
                # Determine 'gelar'
                gender = row.get('gender_1', '')
                status = row.get('marital_1', '')
                if gender == 'male':
                    gelar = 'Tn'
                elif gender == 'female' and status == 'single':
                    gelar = 'Nn'
                elif gender == 'female' and status == 'married':
                    gelar = 'Ny'
                else:
                    gelar = ''  # Default or handle other cases

                # Skip if nomor_urut already exists
                if current_nomor_urut in existing_nomor_urut:
                    current_nomor_urut += 1  # Increment nomor_urut to avoid duplicates
                
                record = Record(
                    nomor_urut=current_nomor_urut,
                    nomor_akta=nomor_akta,
                    tanggal_akta=tanggal_akta,
                    sifat_akta='AKTA JAMINAN FIDUSIA',
                    nama_debitur=row.get('name_debitur', ''),
                    gender=row.get('gender_1', ''),
                    status=row.get('marital_1', ''),
                    nama_perwakilan=nama_perwakilan,
                    perusahaan_finance=nama_perusahaan_finance,
                    alamat_perusahaan=alamat_perusahaan_finance,
                    gelar=gelar
                )
                records.append(record)
                current_nomor_urut += 1  # Increment nomor_urut for the next record
            
            # Add records to the session and commit
            session.add_all(records)
            session.commit()
            
            messagebox.showinfo("Berhasil", "Data berhasil di simpan ke database!")
            self.upload_window.destroy()  # Close the upload form
            self.display_data()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            self.log(f"Error processing file {file_path}: {e}")

    def display_data(self, search_query=""):
        # Clear the existing data in the Treeview
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        # Query all records from the database, with optional filtering
        # records = session.query(Record).order_by(Record.nomor_urut).all()
        query = session.query(Record).order_by(Record.nomor_urut)
        if search_query:
            search_query = f"%{search_query}%"
            query = query.filter(
                (Record.nomor_akta.like(search_query)) |
                (Record.nama_debitur.like(search_query)) |
                (Record.nama_perwakilan.like(search_query)) |
                (Record.perusahaan_finance.like(search_query)) |
                (Record.alamat_perusahaan.like(search_query))
            )
        records = query.all()
        
        # Insert records into the Treeview
        for index, record in enumerate(records):
            tag = 'evenrow' if index % 2 == 0 else 'oddrow'
            self.tree.insert('', 'end', iid=record.id, values=(
                record.nomor_urut,
                record.nomor_akta,
                record.tanggal_akta,
                record.sifat_akta,
                record.nama_debitur,
                record.gender,
                record.status,
                record.nama_perwakilan,
                record.perusahaan_finance,
                record.alamat_perusahaan,
                record.gelar
            ), tags=(tag,))
            
        self.adjust_column_widths()

    def adjust_column_widths(self):
        for col in self.tree['columns']:
            self.tree.column(col, width=tk.NO)  # Disable auto width
        self.tree.update_idletasks()

    def on_item_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        col_index = int(column.split('#')[-1]) - 1
        current_value = self.tree.item(item_id, 'values')[col_index]
        
        self.edit_window = tk.Toplevel(self.root)
        self.edit_window.title("Edit Data")
        self.edit_window.geometry("200x200")  # Set modal size to 150x150

        tk.Label(self.edit_window, text="Input Data Baru:").pack(pady=10)
        self.edit_entry = tk.Entry(self.edit_window)
        self.edit_entry.pack(pady=5)
        self.edit_entry.insert(0, current_value)
        
        tk.Button(self.edit_window, text="Simpan", command=lambda: self.save_edit(item_id, col_index)).pack(pady=10)

    def save_edit(self, item_id, col_index):
        new_value = self.edit_entry.get()
        
        # Update the Treeview
        self.tree.item(item_id, values=(
            *self.tree.item(item_id, 'values')[:col_index],
            new_value,
            *self.tree.item(item_id, 'values')[col_index + 1:]
        ))
        
        # Update the database
        record_id = item_id
        record = session.query(Record).filter_by(id=record_id).first()
        column_name = self.tree['columns'][col_index]
        setattr(record, column_name.lower().replace(' ', '_'), new_value)
        
        session.commit()
        
        self.edit_window.destroy()
        self.display_data()  # Refresh the Treeview to reflect changes

    def open_input_form(self):
        self.top = tk.Toplevel(self.root)
        self.top.title("Input Data")
        self.top.geometry("400x400")  # Set modal size to 400x400

        self.labels = [
            'Nomor Urut',
            'Nomor Akta',
            'Tanggal Akta (DD/MM/YYYY)',
            'Sifat Akta',
            'Nama Debitur',
            'Gender',
            'Status',
            'Nama Perwakilan',
            'Perusahaan Finance',
            'Alamat Perusahaan'
        ]
        self.entries = {}

        for index, label in enumerate(self.labels):
            tk.Label(self.top, text=label).grid(row=index, column=0, sticky='e', padx=10, pady=5)
            
            # if label == 'Tanggal Akta (DD/MM/YYYY)':
            #     # Add DateEntry widget for date selection
            #     entry = DateEntry(self.top, date_pattern='dd/mm/yyyy')
            if label == 'Tanggal Akta':
                # Add Calendar widget for date selection
                entry = Calendar(self.top, selectmode='day', date_pattern='dd/mm/yyyy')
                entry.grid(row=index, column=1, padx=10, pady=5)
                self.entries[label] = entry
            elif label == 'Gender':
                # Add Combobox for gender selection
                entry = ttk.Combobox(self.top, values=['male', 'female'])
            elif label == 'Status':
                # Add Combobox for status selection
                entry = ttk.Combobox(self.top, values=['single', 'married'])
            else:
                # Regular Entry widget for other inputs
                entry = tk.Entry(self.top)
                
            entry.grid(row=index, column=1, padx=10, pady=5)
            self.entries[label] = entry

        tk.Button(self.top, text="Simpan", command=self.save_input_data).grid(row=len(self.labels), column=0, columnspan=2, pady=10)
        
    def save_input_data(self):
        try:
            nomor_akta = self.entries['Nomor Akta'].get()
            tanggal_akta_str = self.entries['Tanggal Akta (DD/MM/YYYY)'].get()
            sifat_akta = self.entries['Sifat Akta'].get()
            nama_debitur = self.entries['Nama Debitur'].get()
            gender = self.entries['Gender'].get()
            status = self.entries['Status'].get()
            nama_perwakilan = self.entries['Nama Perwakilan'].get()
            perusahaan_finance = self.entries['Perusahaan Finance'].get()
            alamat_perusahaan = self.entries['Alamat Perusahaan'].get()
            nomor_urut_baru = int(self.entries["Nomor Urut"].get())
            
            # Set Gelar based on Gender and Status
            if gender == 'male':
                gelar = 'Tn'
            elif gender == 'female' and status == 'single':
                gelar = 'Nn'
            elif gender == 'female' and status == 'married':
                gelar = 'Ny'
            else:
                gelar = ''

            # Convert date from DD/MM/YYYY to DD Month YYYY
            try:
                tanggal_akta = datetime.datetime.strptime(tanggal_akta_str, "%d/%m/%Y").strftime("%d %B %Y")
            except ValueError:
                tanggal_akta = ''  # Default to empty if format is incorrect

            # Cek apakah nomor urut baru sudah ada
            existing_record = session.query(Record).filter_by(nomor_urut=nomor_urut_baru).first()

            if existing_record:
                # Jika nomor urut baru sudah ada, geser nomor urut yang ada
                records_to_update = session.query(Record).filter(Record.nomor_urut >= nomor_urut_baru).all()
                for record in records_to_update:
                    record.nomor_urut += 1
                
                new_record = Record(
                    nomor_urut=nomor_urut_baru,
                    nomor_akta=nomor_akta,
                    tanggal_akta=tanggal_akta,
                    sifat_akta=sifat_akta,
                    nama_debitur=nama_debitur,
                    gender=gender,
                    status=status,
                    nama_perwakilan=nama_perwakilan,
                    perusahaan_finance=perusahaan_finance,
                    alamat_perusahaan=alamat_perusahaan,
                    gelar=gelar
                )

                # Add record to the session and commit
                session.add(new_record)
                
            session.commit()

            self.log(f"Data saved: {nomor_akta}, {tanggal_akta}, {nama_debitur}, {nomor_urut_baru}")
            messagebox.showinfo("Berhasil","Data berhasil disimpan!")
            self.top.destroy()  # Close the input form
            self.display_data()  # Refresh the Treeview
        except Exception as e:
            session.rollback()  # Rollback jika terjadi error
            messagebox.showerror("Error", f"An error occurred: {e}")
            self.log(f"Error saving data: {e}")


    def shift_nomor_urut(self, new_nomor_urut):
        try:
            # Ambil semua record yang nomor urutnya lebih besar atau sama dengan new_nomor_urut
            records_to_shift = session.query(Record).filter(Record.nomor_urut >= new_nomor_urut).order_by(Record.nomor_urut.desc()).all()

            # Mulai dari nomor urut terbesar, geser semua nomor urut
            for record in records_to_shift:
                record.nomor_urut += 1
                
            session.commit()

            self.log(f"Shifted nomor urut starting from {new_nomor_urut}to {new_nomor_urut + len(records_to_shift)}")
        except Exception as e:
            session.rollback()  # Rollback jika terjadi error
            messagebox.showerror("Error", f"An error occurred while shifting nomor_urut: {e}")
            self.log(f"Error shifting nomor_urut starting from {new_nomor_urut}: {e}")

    def get_next_nomor_urut(self):
        # Get the next nomor_urut based on the existing records
        last_record = session.query(Record).order_by(Record.nomor_urut.desc()).first()
        return (last_record.nomor_urut + 1) if last_record else 1

    def delete_data(self):
        selected_item = self.tree.selection()       
        for item in selected_item:
            record_values = self.tree.item(item, 'values')
            nomor_urut = record_values[0]

            # Delete from database
            record_to_delete = session.query(Record).filter_by(nomor_urut=nomor_urut).first()
            if record_to_delete:
               session.delete(record_to_delete)
               session.commit()

            # Remove from Treeview
            self.tree.delete(item)
            self.display_data()

        messagebox.showinfo("Info", "Record(s) data berhasil dihapus.")
        self.log(f"Records deleted: {selected_item}")

    def download_data(self):
        try:
            # Query all records from the database
            records = session.query(Record).order_by(Record.nomor_urut).all()
            
            # Create a new Excel workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            
            # Set headers in the first row
            ws['A1'] = 'Nomor Urut'
            ws['B1'] = 'Nomor Akta'
            ws['C1'] = 'Tanggal Akta'
            ws['D1'] = 'Sifat Akta'
            ws['E1'] = 'Nama Penghadap dan atau yang diwakilkan/kuasa'
            ws['E2'] = 'Nama Debitur'
            ws['E3'] = 'Perusahaan Finance'
            ws['E4'] = 'Alamat Perusahaan'
            ws['E5'] = 'gelar'
            
            # Apply bold font for headers
            bold_font = Font(bold=True)
            ws['A1'].font = bold_font
            ws['B1'].font = bold_font
            ws['C1'].font = bold_font
            ws['D1'].font = bold_font
            ws['E1'].font = bold_font

            # Set the rest of the font to regular
            regular_font = Font(bold=False)
            ws['E2'].font = regular_font
            ws['E3'].font = regular_font
            ws['E4'].font = regular_font
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply border to header cells
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
                ws[cell].border = thin_border

            # Write data to Excel
            row_index = 2  # Start writing data from the second row
            for record in records:
                # Convert date from YYYY-MM-DD to DD Month YYYY
                tanggal_akta = record.tanggal_akta
                if tanggal_akta:
                    tanggal_akta = datetime.datetime.strptime(tanggal_akta, "%d %B %Y").strftime("%d %B %Y")
                else:
                    tanggal_akta = ''

                ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + 3, end_column=1)
                detail_cell = ws.cell(row=row_index, column=1)
                detail_cell.value = (record.nomor_urut)
                for i in range(row_index, row_index + 4):  # iterasi melalui baris untuk menambahkan border ke setiap sel dalam kolom
                    ws.cell(row=i, column=1).border = thin_border
                
                ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index + 3, end_column=2)
                detail_cell = ws.cell(row=row_index, column=2)
                detail_cell.value = (record.nomor_akta)
                for i in range(row_index, row_index + 4):  # iterasi melalui baris untuk menambahkan border ke setiap sel dalam kolom
                    ws.cell(row=i, column=2).border = thin_border

                ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index + 3, end_column=3)
                detail_cell = ws.cell(row=row_index, column=3)
                detail_cell.value = (record.tanggal_akta)
                for i in range(row_index, row_index + 4):  # iterasi melalui baris untuk menambahkan border ke setiap sel dalam kolom
                    ws.cell(row=i, column=3).border = thin_border

                ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index + 3, end_column=4)
                detail_cell = ws.cell(row=row_index, column=4)
                detail_cell.value = (record.sifat_akta)
                for i in range(row_index, row_index + 4):  # iterasi melalui baris untuk menambahkan border ke setiap sel dalam kolom
                    ws.cell(row=i, column=4).border = thin_border
                

                # ws.cell(row=row_index, column=2, value=record.nomor_akta).border = thin_border
                # ws.cell(row=row_index, column=3, value=record.tanggal_akta).border = thin_border
                # ws.cell(row=row_index, column=4, value=record.sifat_akta).border = thin_border
                
                # Merge cells for the detailed information
                ws.merge_cells(start_row=row_index, start_column=5, end_row=row_index + 3, end_column=5)
                detail_cell = ws.cell(row=row_index, column=5)
                detail_cell.value = (
                    f"- {record.nama_perwakilan}\n"
                    f"QQ.   A. {record.gelar}. {record.nama_debitur}\n"
                    f"           B. PT. {record.perusahaan_finance}\n"
                    f"                Berkedudukan di {record.alamat_perusahaan}"
                )
                detail_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Apply border to the merged cell area
                for row in range(row_index, row_index + 4):
                    ws.cell(row=row, column=5).border = thin_border

                # Move to the next set of rows
                row_index += 4
                
            
            # Save the workbook
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Berhasil", "Data Excel berhasil penyesuaian template dan disimpan!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            self.log(f"Data downloaded to file: {file_path}")
            
    def filter_data(self):
        search_query = self.search_entry.get()
        self.display_data(search_query)

    # Fungsi untuk menghapus semua data di database
    def delete_all_data(self):
        confirm = messagebox.askyesno("Konfirmasi", "Apakah Anda yakin ingin menghapus semua data?")
        if confirm:
            try:
                session.query(Record).delete()
                session.commit()
                self.display_data()  # Refresh Treeview setelah penghapusan
                messagebox.showinfo("Berhasil", "Semua data berhasil dihapus.")
                self.log("All records deleted from the database.")
            except Exception as e:
                session.rollback()  # Rollback jika terjadi error
                messagebox.showerror("Error", f"An error occurred: {e}")
                self.log(f"Error deleting all data: {e}")

        

if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()
